"""
Pilgrim Import API Views for Admin Dashboard.

Provides functionality to:
1. Download a template Excel file for bulk pilgrim import
2. Upload and process Excel file to import pilgrims
"""

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io

from apps.accounts.models import Account, PilgrimProfile
from apps.common.permissions import IsStaff


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsStaff])
def download_template(request):
    """
    Download an Excel template for bulk pilgrim import.
    
    Returns an Excel file with:
    - Properly formatted headers
    - Data validation hints
    - Example row
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Pilgrims"
    
    # Define headers
    headers = [
        'Full Name*',
        'Passport Number*',
        'Phone Number*',
        'Date of Birth (YYYY-MM-DD)',
        'Gender (MALE/FEMALE/OTHER)',
        'Nationality (2-letter code)',
        'Address',
        'Emergency Contact Name',
        'Emergency Contact Phone',
        'Emergency Relationship',
        'Medical Conditions',
    ]
    
    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add example row
    example_row = [
        'Ahmed Ibrahim Hassan',
        'P12345678',
        '+256700000000',
        '1990-05-15',
        'MALE',
        'UG',
        'Kampala, Uganda',
        'Fatima Hassan',
        '+256700000001',
        'Wife',
        'None',
    ]
    
    for col_num, value in enumerate(example_row, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Add instructions worksheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["Pilgrim Import Instructions", ""],
        ["", ""],
        ["Required Fields (marked with *):", ""],
        ["- Full Name", "Full name as on passport"],
        ["- Passport Number", "Must be unique, will be used for verification"],
        ["- Phone Number", "Must be unique with country code (e.g., +256700000000)"],
        ["", ""],
        ["Date Format:", ""],
        ["- Date of Birth", "Use YYYY-MM-DD format (e.g., 1990-05-15)"],
        ["", ""],
        ["Gender Options:", ""],
        ["", "MALE, FEMALE, or OTHER"],
        ["", ""],
        ["Nationality:", ""],
        ["", "Use ISO 3166-1 alpha-2 country codes"],
        ["", "Examples: UG (Uganda), KE (Kenya), TZ (Tanzania), NG (Nigeria)"],
        ["", ""],
        ["Notes:", ""],
        ["- Delete the example row before uploading", ""],
        ["- Ensure phone numbers include country code", ""],
        ["- Passport numbers must be unique", ""],
        ["- Leave optional fields empty if not available", ""],
    ]
    
    # Style instructions
    title_font = Font(bold=True, size=14, color="366092")
    section_font = Font(bold=True, size=11)
    
    for row_num, (col1, col2) in enumerate(instructions, 1):
        ws_instructions.cell(row=row_num, column=1).value = col1
        ws_instructions.cell(row=row_num, column=2).value = col2
        
        if row_num == 1:
            ws_instructions.cell(row=row_num, column=1).font = title_font
        elif col1 and col1.endswith(":"):
            ws_instructions.cell(row=row_num, column=1).font = section_font
    
    # Set column widths
    ws.column_dimensions['A'].width = 25  # Full Name
    ws.column_dimensions['B'].width = 18  # Passport Number
    ws.column_dimensions['C'].width = 18  # Phone
    ws.column_dimensions['D'].width = 22  # DOB
    ws.column_dimensions['E'].width = 28  # Gender
    ws.column_dimensions['F'].width = 28  # Nationality
    ws.column_dimensions['G'].width = 30  # Address
    ws.column_dimensions['H'].width = 25  # Emergency Name
    ws.column_dimensions['I'].width = 20  # Emergency Phone
    ws.column_dimensions['J'].width = 20  # Emergency Relationship
    ws.column_dimensions['K'].width = 30  # Medical Conditions
    
    ws_instructions.column_dimensions['A'].width = 30
    ws_instructions.column_dimensions['B'].width = 50
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=pilgrims_import_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStaff])
def validate_import(request):
    """
    Validate pilgrim import file and check for duplicates.
    
    This is step 1 of a 2-phase import process:
    1. Validate & Preview - Parse file, check duplicates, return preview
    2. Confirm Import - Actually create the records
    
    Returns:
    - valid_rows: Rows that can be imported
    - duplicate_rows: Rows that match existing pilgrims
    - error_rows: Rows with validation errors
    """
    if 'file' not in request.FILES:
        return Response(
            {'error': 'No file provided'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    file = request.FILES['file']
    
    # Validate file type
    if not file.name.endswith('.xlsx'):
        return Response(
            {'error': 'Invalid file type. Please upload an Excel file (.xlsx)'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Load workbook
        wb = load_workbook(file, read_only=False, data_only=True)
        ws = wb.active
        
        # Validate headers
        expected_headers = [
            'Full Name*',
            'Passport Number*',
            'Phone Number*',
            'Date of Birth (YYYY-MM-DD)',
            'Gender (MALE/FEMALE/OTHER)',
            'Nationality (2-letter code)',
            'Address',
            'Emergency Contact Name',
            'Emergency Contact Phone',
            'Emergency Relationship',
            'Medical Conditions',
        ]
        
        actual_headers = [cell.value for cell in ws[1]]
        
        for i, expected in enumerate(expected_headers):
            if i >= len(actual_headers) or actual_headers[i] != expected:
                return Response(
                    {
                        'error': f'Invalid template format. Expected header "{expected}" at column {i+1}',
                        'hint': 'Please download the latest template and try again'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        valid_rows = []
        duplicate_rows = []
        error_rows = []
        
        # Get existing phone numbers and passport numbers for duplicate checking
        existing_phones = set(Account.objects.filter(role='PILGRIM').values_list('phone', flat=True))
        existing_passports = set(PilgrimProfile.objects.exclude(passport_number='').values_list('passport_number', flat=True))
        
        # Process each row (skip header row)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue
            
            # Extract fields
            full_name = row[0]
            passport_number = row[1]
            phone = row[2]
            dob_str = row[3]
            gender_str = row[4]
            nationality = row[5]
            address = row[6]
            emergency_name = row[7]
            emergency_phone = row[8]
            emergency_relationship = row[9]
            medical_conditions = row[10]
            
            row_errors = []
            row_warnings = []
            duplicate_type = None
            
            # Validate required fields
            if not full_name or str(full_name).strip() == '':
                row_errors.append('Full Name is required')
            
            if not passport_number or str(passport_number).strip() == '':
                row_errors.append('Passport Number is required')
            else:
                passport_number = str(passport_number).strip()
            
            if not phone or str(phone).strip() == '':
                row_errors.append('Phone Number is required')
            else:
                phone = str(phone).strip()
                # Basic phone validation
                if not phone.startswith('+'):
                    row_warnings.append('Phone should include country code (e.g., +256...)')
            
            # Parse date of birth
            dob = None
            if dob_str:
                try:
                    if isinstance(dob_str, datetime):
                        dob = dob_str.date()
                    else:
                        dob = datetime.strptime(str(dob_str), '%Y-%m-%d').date()
                except ValueError:
                    row_errors.append(f'Invalid date format: {dob_str}. Expected YYYY-MM-DD')
            
            # Validate gender
            gender = None
            if gender_str:
                gender_upper = str(gender_str).strip().upper()
                if gender_upper not in ['MALE', 'FEMALE', 'OTHER']:
                    row_errors.append(f'Invalid gender: {gender_str}. Must be MALE, FEMALE, or OTHER')
                else:
                    gender = gender_upper
            
            # Validate nationality (2-letter code)
            if nationality and len(str(nationality).strip()) != 2:
                row_warnings.append(f'Nationality should be 2-letter country code, got: {nationality}')
            
            # Check for duplicates
            if phone and phone in existing_phones:
                duplicate_type = 'phone'
                row_warnings.append(f'Phone number {phone} already exists')
            
            if passport_number and passport_number in existing_passports:
                if duplicate_type:
                    duplicate_type = 'both'
                else:
                    duplicate_type = 'passport'
                row_warnings.append(f'Passport number {passport_number} already exists')
            
            # Build row data
            row_data = {
                'row_number': row_num,
                'full_name': full_name,
                'passport_number': passport_number,
                'phone': phone,
                'dob': str(dob) if dob else None,
                'gender': gender,
                'nationality': nationality,
                'address': address,
                'emergency_name': emergency_name,
                'emergency_phone': emergency_phone,
                'emergency_relationship': emergency_relationship,
                'medical_conditions': medical_conditions,
                'errors': row_errors,
                'warnings': row_warnings,
                'duplicate_type': duplicate_type,
            }
            
            # Categorize row
            if row_errors:
                error_rows.append(row_data)
            elif duplicate_type:
                duplicate_rows.append(row_data)
            else:
                valid_rows.append(row_data)
        
        # Prepare response
        response_data = {
            'success': True,
            'summary': {
                'total': len(valid_rows) + len(duplicate_rows) + len(error_rows),
                'valid': len(valid_rows),
                'duplicates': len(duplicate_rows),
                'errors': len(error_rows),
            },
            'valid_rows': valid_rows,
            'duplicate_rows': duplicate_rows,
            'error_rows': error_rows,
            'message': f'Found {len(valid_rows)} valid row(s), {len(duplicate_rows)} duplicate(s), {len(error_rows)} error(s)'
        }
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to process file: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStaff])
def import_pilgrims(request):
    """
    Upload and process Excel file to import pilgrims.
    
    Expected file format:
    - Excel file (.xlsx)
    - Headers in first row matching template
    - Data starting from row 2
    
    Returns:
    - success: True/False
    - imported: Number of successfully imported pilgrims
    - errors: List of errors with row numbers
    - warnings: List of warnings
    """
    if 'file' not in request.FILES:
        return Response(
            {'error': 'No file provided'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    file = request.FILES['file']
    
    # Validate file type
    if not file.name.endswith('.xlsx'):
        return Response(
            {'error': 'Invalid file type. Please upload an Excel file (.xlsx)'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # Load workbook
        wb = load_workbook(file, read_only=False, data_only=True)
        ws = wb.active
        
        # Validate headers (first row)
        expected_headers = [
            'Full Name*',
            'Passport Number*',
            'Phone Number*',
            'Date of Birth (YYYY-MM-DD)',
            'Gender (MALE/FEMALE/OTHER)',
            'Nationality (2-letter code)',
            'Address',
            'Emergency Contact Name',
            'Emergency Contact Phone',
            'Emergency Relationship',
            'Medical Conditions',
        ]
        
        actual_headers = [cell.value for cell in ws[1]]
        
        # Check if headers match (allowing for extra columns)
        for i, expected in enumerate(expected_headers):
            if i >= len(actual_headers) or actual_headers[i] != expected:
                return Response(
                    {
                        'error': f'Invalid template format. Expected header "{expected}" at column {i+1}',
                        'hint': 'Please download the latest template and try again'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Process rows
        imported_count = 0
        errors = []
        warnings = []
        
        # Start transaction
        with transaction.atomic():
            for row_num in range(2, ws.max_row + 1):
                row = ws[row_num]
                
                # Skip empty rows
                if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                    continue
                
                try:
                    # Extract data
                    full_name = str(row[0].value).strip() if row[0].value else None
                    passport_number = str(row[1].value).strip() if row[1].value else None
                    phone = str(row[2].value).strip() if row[2].value else None
                    dob_value = row[3].value
                    gender = str(row[4].value).strip().upper() if row[4].value else None
                    nationality = str(row[5].value).strip().upper() if row[5].value else None
                    address = str(row[6].value).strip() if row[6].value else None
                    emergency_name = str(row[7].value).strip() if row[7].value else None
                    emergency_phone = str(row[8].value).strip() if row[8].value else None
                    emergency_relationship = str(row[9].value).strip() if row[9].value else None
                    medical_conditions = str(row[10].value).strip() if row[10].value else None
                    
                    # Validate required fields
                    if not full_name:
                        errors.append(f"Row {row_num}: Full Name is required")
                        continue
                    
                    if not passport_number:
                        errors.append(f"Row {row_num}: Passport Number is required")
                        continue
                    
                    if not phone:
                        errors.append(f"Row {row_num}: Phone Number is required")
                        continue
                    
                    # Validate and parse date of birth
                    dob = None
                    if dob_value:
                        try:
                            if isinstance(dob_value, datetime):
                                dob = dob_value.date()
                            elif isinstance(dob_value, str):
                                dob = datetime.strptime(dob_value, '%Y-%m-%d').date()
                        except ValueError:
                            errors.append(f"Row {row_num}: Invalid date format for Date of Birth. Use YYYY-MM-DD")
                            continue
                    
                    # Validate gender
                    if gender and gender not in ['MALE', 'FEMALE', 'OTHER']:
                        errors.append(f"Row {row_num}: Gender must be MALE, FEMALE, or OTHER")
                        continue
                    
                    # Validate nationality (2-letter code)
                    if nationality and len(nationality) != 2:
                        errors.append(f"Row {row_num}: Nationality must be a 2-letter ISO country code")
                        continue
                    
                    # Check for duplicate passport number
                    if PilgrimProfile.objects.filter(passport_number=passport_number).exists():
                        errors.append(f"Row {row_num}: Passport Number '{passport_number}' already exists")
                        continue
                    
                    # Check for duplicate phone
                    if PilgrimProfile.objects.filter(phone=phone).exists():
                        errors.append(f"Row {row_num}: Phone Number '{phone}' already exists")
                        continue
                    
                    # Check if account with this phone exists
                    if Account.objects.filter(phone=phone).exists():
                        errors.append(f"Row {row_num}: An account with phone '{phone}' already exists")
                        continue
                    
                    # Create Account
                    account = Account.objects.create_user(
                        phone=phone,
                        name=full_name,
                        role='PILGRIM',
                        is_staff=False
                    )
                    
                    # Create PilgrimProfile
                    pilgrim = PilgrimProfile.objects.create(
                        user=account,
                        full_name=full_name,
                        passport_number=passport_number,
                        phone=phone,
                        dob=dob,
                        gender=gender,
                        nationality=nationality,
                        address=address or '',
                        emergency_name=emergency_name or '',
                        emergency_phone=emergency_phone or '',
                        emergency_relationship=emergency_relationship or '',
                        medical_conditions=medical_conditions or '',
                        created_by=request.user
                    )
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    continue
        
        # Prepare response
        response_data = {
            'success': imported_count > 0,
            'imported': imported_count,
            'errors': errors,
            'warnings': warnings,
            'message': f'Successfully imported {imported_count} pilgrim(s)'
        }
        
        if errors:
            response_data['message'] = f'Imported {imported_count} pilgrim(s) with {len(errors)} error(s)'
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': f'Failed to process file: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )

