"""
Unit tests for Pilgrim Import API endpoints.
"""

from django.test import TestCase
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework import status
from openpyxl import Workbook
import io

from apps.accounts.models import Account, PilgrimProfile


class PilgrimImportTests(TestCase):
    """Test suite for pilgrim import functionality."""
    
    def setUp(self):
        """Set up test client and create test user."""
        self.client = APIClient()
        
        # Create staff user
        self.staff_user = Account.objects.create_user(
            phone='+256700000000',
            name='Test Staff',
            role='STAFF',
            is_staff=True
        )
        
        # Authenticate
        self.client.force_authenticate(user=self.staff_user)
    
    def test_download_template_success(self):
        """Test successful template download."""
        url = reverse('api:pilgrim-import-template')
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('pilgrims_import_template', response['Content-Disposition'])
    
    def test_download_template_requires_auth(self):
        """Test that template download requires authentication."""
        self.client.force_authenticate(user=None)
        url = reverse('api:pilgrim-import-template')
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
    
    def test_download_template_requires_staff(self):
        """Test that template download requires staff role."""
        # Create pilgrim user
        pilgrim_user = Account.objects.create_user(
            phone='+256700000001',
            name='Test Pilgrim',
            role='PILGRIM',
            is_staff=False
        )
        self.client.force_authenticate(user=pilgrim_user)
        
        url = reverse('api:pilgrim-import-template')
        response = self.client.get(url)
        
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
    
    def test_import_pilgrims_no_file(self):
        """Test import with no file provided."""
        url = reverse('api:pilgrim-import')
        response = self.client.post(url, {})
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)
        self.assertIn('No file provided', response.data['error'])
    
    def test_import_pilgrims_invalid_file_type(self):
        """Test import with invalid file type."""
        url = reverse('api:pilgrim-import')
        
        # Create a text file
        file_content = b'This is a text file'
        file = io.BytesIO(file_content)
        file.name = 'test.txt'
        
        response = self.client.post(url, {'file': file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)
        self.assertIn('Invalid file type', response.data['error'])
    
    def test_import_pilgrims_success(self):
        """Test successful pilgrim import."""
        url = reverse('api:pilgrim-import')
        
        # Create valid Excel file
        wb = Workbook()
        ws = wb.active
        
        # Headers
        headers = [
            'Full Name*',
            'Passport Number*',
            'Phone Number*',
            'Date of Birth (YYYY-MM-DD)',
            'Gender (MALE/FEMALE/OTHER)',
            'Nationality (2-letter code)',
            'Address',
            'Emergency Contact Name',
            'Emergency Contact Phone',
            'Emergency Relationship',
            'Medical Conditions',
        ]
        ws.append(headers)
        
        # Data row
        ws.append([
            'Ahmed Ibrahim',
            'P12345678',
            '+256700111111',
            '1990-05-15',
            'MALE',
            'UG',
            'Kampala, Uganda',
            'Fatima Hassan',
            '+256700111112',
            'Wife',
            'None',
        ])
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'pilgrims.xlsx'
        
        response = self.client.post(url, {'file': excel_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['imported'], 1)
        self.assertEqual(len(response.data['errors']), 0)
        
        # Verify pilgrim was created
        self.assertTrue(PilgrimProfile.objects.filter(passport_number='P12345678').exists())
        pilgrim = PilgrimProfile.objects.get(passport_number='P12345678')
        self.assertEqual(pilgrim.full_name, 'Ahmed Ibrahim')
        self.assertEqual(pilgrim.phone, '+256700111111')
        self.assertEqual(pilgrim.gender, 'MALE')
        self.assertEqual(pilgrim.nationality, 'UG')
    
    def test_import_pilgrims_missing_required_fields(self):
        """Test import with missing required fields."""
        url = reverse('api:pilgrim-import')
        
        # Create Excel file with missing required field
        wb = Workbook()
        ws = wb.active
        
        # Headers
        headers = [
            'Full Name*',
            'Passport Number*',
            'Phone Number*',
            'Date of Birth (YYYY-MM-DD)',
            'Gender (MALE/FEMALE/OTHER)',
            'Nationality (2-letter code)',
            'Address',
            'Emergency Contact Name',
            'Emergency Contact Phone',
            'Emergency Relationship',
            'Medical Conditions',
        ]
        ws.append(headers)
        
        # Data row with missing phone
        ws.append([
            'Ahmed Ibrahim',
            'P12345678',
            '',  # Missing phone
            '1990-05-15',
            'MALE',
            'UG',
            'Kampala, Uganda',
            '',
            '',
            '',
            '',
        ])
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'pilgrims.xlsx'
        
        response = self.client.post(url, {'file': excel_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['imported'], 0)
        self.assertGreater(len(response.data['errors']), 0)
        self.assertIn('Phone Number is required', response.data['errors'][0])
    
    def test_import_pilgrims_duplicate_passport(self):
        """Test import with duplicate passport number."""
        # Create existing pilgrim
        existing_account = Account.objects.create_user(
            phone='+256700222222',
            name='Existing Pilgrim',
            role='PILGRIM'
        )
        PilgrimProfile.objects.create(
            user=existing_account,
            full_name='Existing Pilgrim',
            passport_number='P99999999',
            phone='+256700222222'
        )
        
        url = reverse('api:pilgrim-import')
        
        # Create Excel file with duplicate passport
        wb = Workbook()
        ws = wb.active
        
        # Headers
        headers = [
            'Full Name*',
            'Passport Number*',
            'Phone Number*',
            'Date of Birth (YYYY-MM-DD)',
            'Gender (MALE/FEMALE/OTHER)',
            'Nationality (2-letter code)',
            'Address',
            'Emergency Contact Name',
            'Emergency Contact Phone',
            'Emergency Relationship',
            'Medical Conditions',
        ]
        ws.append(headers)
        
        # Data row with duplicate passport
        ws.append([
            'New Pilgrim',
            'P99999999',  # Duplicate
            '+256700333333',
            '1990-05-15',
            'MALE',
            'UG',
            '',
            '',
            '',
            '',
            '',
        ])
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'pilgrims.xlsx'
        
        response = self.client.post(url, {'file': excel_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['imported'], 0)
        self.assertGreater(len(response.data['errors']), 0)
        self.assertIn('already exists', response.data['errors'][0])
    
    def test_import_pilgrims_invalid_date_format(self):
        """Test import with invalid date format."""
        url = reverse('api:pilgrim-import')
        
        # Create Excel file with invalid date
        wb = Workbook()
        ws = wb.active
        
        # Headers
        headers = [
            'Full Name*',
            'Passport Number*',
            'Phone Number*',
            'Date of Birth (YYYY-MM-DD)',
            'Gender (MALE/FEMALE/OTHER)',
            'Nationality (2-letter code)',
            'Address',
            'Emergency Contact Name',
            'Emergency Contact Phone',
            'Emergency Relationship',
            'Medical Conditions',
        ]
        ws.append(headers)
        
        # Data row with invalid date format
        ws.append([
            'Ahmed Ibrahim',
            'P12345678',
            '+256700444444',
            '15/05/1990',  # Invalid format
            'MALE',
            'UG',
            '',
            '',
            '',
            '',
            '',
        ])
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'pilgrims.xlsx'
        
        response = self.client.post(url, {'file': excel_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['imported'], 0)
        self.assertGreater(len(response.data['errors']), 0)
        self.assertIn('Invalid date format', response.data['errors'][0])
    
    def test_import_pilgrims_invalid_gender(self):
        """Test import with invalid gender value."""
        url = reverse('api:pilgrim-import')
        
        # Create Excel file with invalid gender
        wb = Workbook()
        ws = wb.active
        
        # Headers
        headers = [
            'Full Name*',
            'Passport Number*',
            'Phone Number*',
            'Date of Birth (YYYY-MM-DD)',
            'Gender (MALE/FEMALE/OTHER)',
            'Nationality (2-letter code)',
            'Address',
            'Emergency Contact Name',
            'Emergency Contact Phone',
            'Emergency Relationship',
            'Medical Conditions',
        ]
        ws.append(headers)
        
        # Data row with invalid gender
        ws.append([
            'Ahmed Ibrahim',
            'P12345678',
            '+256700555555',
            '1990-05-15',
            'INVALID',  # Invalid gender
            'UG',
            '',
            '',
            '',
            '',
            '',
        ])
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'pilgrims.xlsx'
        
        response = self.client.post(url, {'file': excel_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['imported'], 0)
        self.assertGreater(len(response.data['errors']), 0)
        self.assertIn('Gender must be', response.data['errors'][0])
    
    def test_import_pilgrims_multiple_rows(self):
        """Test import with multiple valid rows."""
        url = reverse('api:pilgrim-import')
        
        # Create Excel file with multiple rows
        wb = Workbook()
        ws = wb.active
        
        # Headers
        headers = [
            'Full Name*',
            'Passport Number*',
            'Phone Number*',
            'Date of Birth (YYYY-MM-DD)',
            'Gender (MALE/FEMALE/OTHER)',
            'Nationality (2-letter code)',
            'Address',
            'Emergency Contact Name',
            'Emergency Contact Phone',
            'Emergency Relationship',
            'Medical Conditions',
        ]
        ws.append(headers)
        
        # Multiple data rows
        ws.append(['Ahmed Ibrahim', 'P11111111', '+256700666666', '1990-05-15', 'MALE', 'UG', '', '', '', '', ''])
        ws.append(['Fatima Hassan', 'P22222222', '+256700666667', '1992-08-20', 'FEMALE', 'KE', '', '', '', '', ''])
        ws.append(['Omar Ali', 'P33333333', '+256700666668', '1988-12-10', 'MALE', 'TZ', '', '', '', '', ''])
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'pilgrims.xlsx'
        
        response = self.client.post(url, {'file': excel_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['imported'], 3)
        self.assertEqual(len(response.data['errors']), 0)
        
        # Verify all pilgrims were created
        self.assertEqual(PilgrimProfile.objects.count(), 3)
    
    def test_import_pilgrims_skips_empty_rows(self):
        """Test import skips empty rows."""
        url = reverse('api:pilgrim-import')
        
        # Create Excel file with empty rows
        wb = Workbook()
        ws = wb.active
        
        # Headers
        headers = [
            'Full Name*',
            'Passport Number*',
            'Phone Number*',
            'Date of Birth (YYYY-MM-DD)',
            'Gender (MALE/FEMALE/OTHER)',
            'Nationality (2-letter code)',
            'Address',
            'Emergency Contact Name',
            'Emergency Contact Phone',
            'Emergency Relationship',
            'Medical Conditions',
        ]
        ws.append(headers)
        
        # Data with empty rows
        ws.append(['Ahmed Ibrahim', 'P44444444', '+256700777777', '1990-05-15', 'MALE', 'UG', '', '', '', '', ''])
        ws.append(['', '', '', '', '', '', '', '', '', '', ''])  # Empty row
        ws.append(['Fatima Hassan', 'P55555555', '+256700777778', '1992-08-20', 'FEMALE', 'KE', '', '', '', '', ''])
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'pilgrims.xlsx'
        
        response = self.client.post(url, {'file': excel_file}, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['success'])
        self.assertEqual(response.data['imported'], 2)  # Only 2 valid rows
        self.assertEqual(len(response.data['errors']), 0)

